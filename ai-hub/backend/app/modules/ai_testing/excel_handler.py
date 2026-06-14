"""AI Testing Excel 导入/导出处理

导出：12 列 xlsx（序号/标题/项目/版本/优先级/类型/前置条件/步骤/预期结果/标签/状态/创建时间）
导入：解析 xlsx 并返回用例列表，供 service.batch_create_cases() 使用

优先级着色：P0 红 / P1 橙 / P2 蓝 / P3 灰
"""

import io
import logging
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── 样式常量 ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, color="FFFFFF", bold=True)

PRIORITY_FILLS = {
  "P0": PatternFill(start_color="FF4D4F", end_color="FF4D4F", fill_type="solid"),  # 红
  "P1": PatternFill(start_color="FF7A00", end_color="FF7A00", fill_type="solid"),  # 橙
  "P2": PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid"),  # 蓝
  "P3": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),  # 灰
}

THIN_BORDER = Border(
  left=Side(style="thin", color="D9D9D9"),
  right=Side(style="thin", color="D9D9D9"),
  top=Side(style="thin", color="D9D9D9"),
  bottom=Side(style="thin", color="D9D9D9"),
)

WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

HEADERS = [
  "序号", "标题", "项目", "版本", "优先级",
  "用例类型", "前置条件", "测试步骤", "预期结果",
  "标签", "状态", "创建时间",
]

COL_WIDTHS = [6, 30, 14, 8, 8, 12, 28, 40, 40, 16, 10, 18]


# ── 导出 ──────────────────────────────────────────────────────────────────────
def export_cases_to_xlsx(
  cases: list[dict[str, Any]],
  project_name: str = "",
) -> bytes:
  """将用例列表导出为 xlsx 字节流

  Args:
    cases: 用例列表，每个 dict 含 title/priority/case_type/preconditions/steps/expected_results/tags/status/created_at
    project_name: 项目名称（填入"项目"列）

  Returns:
    xlsx 文件字节内容
  """
  wb = Workbook()
  ws = wb.active
  ws.title = "测试用例"

  # 写表头
  for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

  # 列宽
  for col_idx, width in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

  # 冻结首行
  ws.freeze_panes = "A2"

  # 写数据行
  for row_idx, case in enumerate(cases, 1):
    priority = str(case.get("priority", "")).upper()
    tags = case.get("tags", [])
    if isinstance(tags, list):
      tags_str = ", ".join(str(t) for t in tags)
    else:
      tags_str = str(tags)

    row_data = [
      row_idx,
      case.get("title", ""),
      project_name,
      case.get("version", "v1.0"),
      priority,
      case.get("case_type", ""),
      case.get("preconditions", ""),
      case.get("steps", ""),
      case.get("expected_results", ""),
      tags_str,
      case.get("status", "draft"),
      case.get("created_at", ""),
    ]

    for col_idx, value in enumerate(row_data, 1):
      cell = ws.cell(row=row_idx + 1, column=col_idx, value=value)
      cell.alignment = WRAP_ALIGN
      cell.border = THIN_BORDER

    # 优先级着色（整行）
    fill = PRIORITY_FILLS.get(priority)
    if fill:
      for col_idx in range(1, len(HEADERS) + 1):
        ws.cell(row=row_idx + 1, column=col_idx).fill = fill

  buf = io.BytesIO()
  wb.save(buf)
  return buf.getvalue()


# ── 导入 ──────────────────────────────────────────────────────────────────────
def parse_xlsx_cases(file_bytes: bytes) -> list[dict[str, Any]]:
  """解析 xlsx 文件，返回用例列表

  期望格式与导出一致（12 列），首行为表头。
  缺失列不报错，返回空字符串。

  Args:
    file_bytes: xlsx 文件字节内容

  Returns:
    用例 dict 列表，每个 dict 可直接传给 repo.create_case()
  """
  buf = io.BytesIO(file_bytes)
  wb = load_workbook(buf, read_only=True, data_only=True)
  ws = wb.active

  cases: list[dict[str, Any]] = []
  rows = list(ws.iter_rows(min_row=2, values_only=True))

  for row in rows:
    if not row or all(cell is None for cell in row):
      continue  # 跳过空行

    # 安全取值，idx 使用 1-based 列号（与 Excel 列号一致）
    def _val(idx: int) -> str:
      col = idx - 1
      if col < len(row) and row[col] is not None:
        return str(row[col]).strip()
      return ""

    title = _val(2)
    if not title:
      continue  # 无标题则跳过

    # 解析标签
    tags_str = _val(10)
    tags = [t.strip() for t in tags_str.split(",") if t.strip()] if tags_str else []

    cases.append({
      "title": title,
      "version": _val(4) or "v1.0",
      "priority": _val(5) or "P2",
      "case_type": _val(6) or "functional",
      "preconditions": _val(7),
      "steps": _val(8),
      "expected_results": _val(9),
      "tags": tags,
      "status": _val(11) or "draft",
    })

  wb.close()
  logger.info(f"[parse_xlsx_cases] 解析到 {len(cases)} 条用例")
  return cases


# ── AI 生成内容解析 ──────────────────────────────────────────────────────────────
def parse_markdown_cases(markdown_text: str) -> list[dict[str, Any]]:
  """解析 AI 生成的 markdown 测试用例为结构化列表

  从 AI 输出（write/revise 阶段）中提取每条用例的结构化字段。
  按 --- 分隔符分割段落后，用正则提取各字段。

  Args:
    markdown_text: AI 输出的 markdown 文本（--- 分隔的多条用例）

  Returns:
    用例 dict 列表，可直接传入 export_cases_to_xlsx() 或 repo.create_case()
  """
  import re

  cases: list[dict[str, Any]] = []

  # 按 --- 分隔符分割段落
  blocks = re.split(r'\n-{3,}\n', markdown_text)

  for block in blocks:
    block = block.strip()
    if not block:
      continue

    # 提取各字段（仅提取 --- 分隔内含 **字段名**: 值的段落）
    def _extract(pattern: str) -> str:
      m = re.search(pattern, block, re.MULTILINE)
      return m.group(1).strip() if m else ""

    title = _extract(r'\*\*标题\*\*\s*:\s*(.+)')
    if not title:
      continue  # 无标题则跳过（可能是分析文本等非用例段落）

    priority_raw = _extract(r'\*\*优先级\*\*\s*:\s*(.+)')
    priority = priority_raw.strip().upper() if priority_raw and priority_raw.strip().upper() in ("P0", "P1", "P2", "P3") else "P2"

    case_type = _extract(r'\*\*用例类型\*\*\s*:\s*(.+)') or "functional"
    # 规范化 AI 输出类型到有效枚举值
    VALID_TYPES = {"functional", "performance", "security", "compatibility", "ui", "api"}
    if case_type not in VALID_TYPES:
      case_type = "functional"
    preconditions = _extract(r'\*\*前置条件\*\*\s*:\s*(.*?)(?=\n\*\*)')
    steps = _extract(r'\*\*测试步骤\*\*\s*:\s*(.*?)(?=\n\*\*)')
    expected_results = _extract(r'\*\*预期结果\*\*\s*:\s*(.*?)(?=\n\*\*)')
    tags_raw = _extract(r'\*\*标签\*\*\s*:\s*(.+)')
    tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

    # 多行字段处理：如果 steps/expected_results 未匹配到（跨行），取 block 内内容
    if not steps:
      steps_match = re.search(r'\*\*测试步骤\*\*\s*:\s*(.+?)(?=\n\*\*|$)', block, re.DOTALL)
      steps = steps_match.group(1).strip() if steps_match else ""
    if not expected_results:
      er_match = re.search(r'\*\*预期结果\*\*\s*:\s*(.+?)(?=\n\*\*|$)', block, re.DOTALL)
      expected_results = er_match.group(1).strip() if er_match else ""

    cases.append({
      "title": title,
      "priority": priority,
      "case_type": case_type,
      "preconditions": preconditions,
      "steps": steps,
      "expected_results": expected_results,
      "tags": tags,
    })

  logger.info(f"[parse_markdown_cases] 从 AI 输出解析到 {len(cases)} 条用例")
  return cases
